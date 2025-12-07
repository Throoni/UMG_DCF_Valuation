"""
Formatting utilities for Excel output and data presentation
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Optional
import numpy as np
import config


def get_header_style():
    """Get style for header cells"""
    return {
        'font': Font(name=config.EXCEL_FORMATTING['font_name'], 
                    size=config.EXCEL_FORMATTING['header_font_size'], 
                    bold=True, color='FFFFFF'),
        'fill': PatternFill(start_color=config.EXCEL_FORMATTING['header_color'],
                           end_color=config.EXCEL_FORMATTING['header_color'],
                           fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }


def get_input_style():
    """Get style for input cells"""
    return {
        'font': Font(name=config.EXCEL_FORMATTING['font_name'],
                    size=config.EXCEL_FORMATTING['data_font_size']),
        'fill': PatternFill(start_color=config.EXCEL_FORMATTING['input_color'],
                           end_color=config.EXCEL_FORMATTING['input_color'],
                           fill_type='solid'),
        'alignment': Alignment(horizontal='right', vertical='center'),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }


def get_calculation_style():
    """Get style for calculation cells"""
    return {
        'font': Font(name=config.EXCEL_FORMATTING['font_name'],
                    size=config.EXCEL_FORMATTING['data_font_size']),
        'fill': PatternFill(start_color=config.EXCEL_FORMATTING['calculation_color'],
                           end_color=config.EXCEL_FORMATTING['calculation_color'],
                           fill_type='solid'),
        'alignment': Alignment(horizontal='right', vertical='center'),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }


def get_formula_style():
    """Get style for formula cells"""
    return {
        'font': Font(name=config.EXCEL_FORMATTING['font_name'],
                    size=config.EXCEL_FORMATTING['data_font_size']),
        'fill': PatternFill(start_color=config.EXCEL_FORMATTING['formula_color'],
                           end_color=config.EXCEL_FORMATTING['formula_color'],
                           fill_type='solid'),
        'alignment': Alignment(horizontal='right', vertical='center'),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }


def apply_style_to_cell(cell, style_dict):
    """Apply style dictionary to a cell"""
    for attr, value in style_dict.items():
        setattr(cell, attr, value)


def format_number(value, format_type: str = 'number') -> str:
    """
    Format number for Excel
    
    Args:
        value: Numeric value
        format_type: 'number', 'percentage', 'currency', 'decimal'
    
    Returns:
        Formatted string
    """
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return ""
    
    if format_type == 'number':
        return config.EXCEL_FORMATTING['number_format']
    elif format_type == 'percentage':
        return config.EXCEL_FORMATTING['percentage_format']
    elif format_type == 'currency':
        return config.EXCEL_FORMATTING['currency_format']
    elif format_type == 'decimal':
        return "0.00"
    else:
        return "General"


def auto_adjust_column_width(worksheet, start_col: int = 1, end_col: Optional[int] = None):
    """
    Auto-adjust column widths in worksheet
    
    Args:
        worksheet: OpenPyXL worksheet object
        start_col: Starting column number (1-indexed)
        end_col: Ending column number (1-indexed), None for all columns
    """
    if end_col is None:
        end_col = worksheet.max_column
    
    for col in range(start_col, end_col + 1):
        max_length = 0
        column_letter = get_column_letter(col)
        
        for cell in worksheet[column_letter]:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        worksheet.column_dimensions[column_letter].width = adjusted_width

