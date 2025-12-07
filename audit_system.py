"""
Comprehensive Audit System for Financial and Technical Validation
Runs after each edit to ensure correctness of calculations and code
"""

import pandas as pd
import numpy as np
from typing import Dict, List, Tuple, Optional
from openpyxl import load_workbook
import os
import config
from utils.data_validation import (
    validate_accounting_identity,
    validate_cash_flow_identity,
    validate_growth_consistency,
    validate_wacc_range,
    validate_terminal_value_pct,
    validate_terminal_growth_rate,
)


class AuditSystem:
    """Comprehensive audit system for financial and technical validation"""
    
    def __init__(self):
        """Initialize audit system"""
        self.audit_results = {
            'financial_checks': [],
            'technical_checks': [],
            'excel_checks': [],
            'errors': [],
            'warnings': [],
        }
    
    def run_full_audit(self, financial_analyzer=None, dcf_model=None,
                      valuation_analyzer=None, excel_file_path=None) -> Dict:
        """
        Run complete audit of the valuation model
        
        Args:
            financial_analyzer: FinancialAnalyzer instance
            dcf_model: DCFModel instance
            valuation_analyzer: ValuationAnalyzer instance
            excel_file_path: Path to Excel output file
        
        Returns:
            Dictionary with audit results
        """
        print("=" * 60)
        print("RUNNING COMPREHENSIVE AUDIT")
        print("=" * 60)
        
        # Financial validation
        if financial_analyzer:
            self._audit_financial_statements(financial_analyzer)
        
        if dcf_model:
            self._audit_dcf_model(dcf_model)
        
        if valuation_analyzer:
            self._audit_valuation_analysis(valuation_analyzer)
        
        # Excel validation
        if excel_file_path and os.path.exists(excel_file_path):
            self._audit_excel_file(excel_file_path)
        
        # Technical validation
        self._audit_code_quality()
        
        # Summary
        self._print_audit_summary()
        
        return self.audit_results
    
    def _audit_financial_statements(self, financial_analyzer):
        """Audit financial statement normalization and ratios"""
        print("\n[Financial Audit] Checking financial statements...")
        
        # Check accounting identity
        if not financial_analyzer.normalized_balance_sheet.empty:
            is_valid, errors = validate_accounting_identity(
                financial_analyzer.normalized_balance_sheet
            )
            if is_valid:
                self.audit_results['financial_checks'].append("✓ Accounting identity validated")
            else:
                self.audit_results['errors'].extend(errors)
                for error in errors:
                    print(f"  ✗ ERROR: {error}")
        
        # Check cash flow identity
        if not financial_analyzer.normalized_cash_flow.empty:
            is_valid, errors = validate_cash_flow_identity(
                financial_analyzer.normalized_cash_flow
            )
            if is_valid:
                self.audit_results['financial_checks'].append("✓ Cash flow identity validated")
            else:
                self.audit_results['errors'].extend(errors)
                for error in errors:
                    print(f"  ✗ ERROR: {error}")
        
        # Check for negative values where inappropriate
        income_stmt = financial_analyzer.normalized_income_stmt
        if not income_stmt.empty:
            if 'Revenue' in income_stmt.columns:
                negative_revenue = (income_stmt['Revenue'] < 0).sum()
                if negative_revenue > 0:
                    error = f"Found {negative_revenue} periods with negative revenue"
                    self.audit_results['errors'].append(error)
                    print(f"  ✗ ERROR: {error}")
                else:
                    self.audit_results['financial_checks'].append("✓ No negative revenue found")
        
        # Check ratio calculations
        if financial_analyzer.ratios:
            if 'gross_margin' in financial_analyzer.ratios:
                margins = financial_analyzer.ratios['gross_margin']
                if any(m < 0 or m > 1 for m in margins if not np.isnan(m)):
                    warning = "Some gross margins are outside 0-100% range"
                    self.audit_results['warnings'].append(warning)
                    print(f"  ⚠ WARNING: {warning}")
                else:
                    self.audit_results['financial_checks'].append("✓ Gross margins are reasonable")
    
    def _audit_dcf_model(self, dcf_model):
        """Audit DCF model calculations"""
        print("\n[DCF Audit] Checking DCF model...")
        
        # Check WACC
        if dcf_model.wacc:
            is_valid, error_msg = validate_wacc_range(
                dcf_model.wacc,
                config.VALIDATION_THRESHOLDS['wacc_min'],
                config.VALIDATION_THRESHOLDS['wacc_max']
            )
            if is_valid:
                self.audit_results['financial_checks'].append(f"✓ WACC is reasonable: {dcf_model.wacc:.2%}")
            else:
                self.audit_results['errors'].append(error_msg)
                print(f"  ✗ ERROR: {error_msg}")
        else:
            self.audit_results['warnings'].append("WACC not calculated")
            print("  ⚠ WARNING: WACC not calculated")
        
        # Check terminal value
        if dcf_model.terminal_value:
            if dcf_model.enterprise_value:
                is_valid, error_msg = validate_terminal_value_pct(
                    dcf_model.terminal_value,
                    dcf_model.enterprise_value,
                    config.VALIDATION_THRESHOLDS['terminal_value_max_pct_of_total']
                )
                if is_valid:
                    tv_pct = dcf_model.terminal_value / dcf_model.enterprise_value
                    self.audit_results['financial_checks'].append(
                        f"✓ Terminal value is reasonable: {tv_pct:.1%} of total value"
                    )
                else:
                    self.audit_results['warnings'].append(error_msg)
                    print(f"  ⚠ WARNING: {error_msg}")
        
        # Check FCFF projections
        if dcf_model.fcff_projections is not None:
            # Check for negative FCFF in terminal year
            final_fcff = dcf_model.fcff_projections.iloc[-1]
            if final_fcff < 0:
                error = f"Final year FCFF is negative: {final_fcff:,.0f}"
                self.audit_results['errors'].append(error)
                print(f"  ✗ ERROR: {error}")
            else:
                self.audit_results['financial_checks'].append(
                    f"✓ Final year FCFF is positive: {final_fcff:,.0f}"
                )
            
            # Check FCFF growth
            if len(dcf_model.fcff_projections) > 1:
                fcff_growth = dcf_model.fcff_projections.pct_change().dropna()
                if any(g < -0.5 for g in fcff_growth if not np.isnan(g)):
                    warning = "Large negative FCFF growth detected"
                    self.audit_results['warnings'].append(warning)
                    print(f"  ⚠ WARNING: {warning}")
        
        # Check revenue projections
        if dcf_model.revenue_projections is not None:
            # Check for negative or zero revenue
            if any(r <= 0 for r in dcf_model.revenue_projections):
                error = "Revenue projections contain non-positive values"
                self.audit_results['errors'].append(error)
                print(f"  ✗ ERROR: {error}")
            else:
                self.audit_results['financial_checks'].append("✓ Revenue projections are positive")
            
            # Check revenue growth reasonableness
            if len(dcf_model.revenue_projections) > 1:
                revenue_growth = dcf_model.revenue_projections.pct_change().dropna()
                if any(g > 0.5 for g in revenue_growth if not np.isnan(g)):
                    warning = "Very high revenue growth (>50%) detected - verify assumptions"
                    self.audit_results['warnings'].append(warning)
                    print(f"  ⚠ WARNING: {warning}")
        
        # Check growth consistency
        if (dcf_model.revenue_projections is not None and 
            dcf_model.income_projections is not None):
            # This would require ROIC calculation - simplified check
            revenue_growth = dcf_model.revenue_projections.pct_change().mean()
            if revenue_growth > 0.20:
                warning = f"High average revenue growth ({revenue_growth:.1%}) - ensure sustainable"
                self.audit_results['warnings'].append(warning)
                print(f"  ⚠ WARNING: {warning}")
    
    def _audit_valuation_analysis(self, valuation_analyzer):
        """Audit valuation analysis"""
        print("\n[Valuation Audit] Checking valuation analysis...")
        
        # Check scenario analysis
        if valuation_analyzer.scenario_results:
            scenarios = ['base', 'bull', 'bear']
            for scenario in scenarios:
                if scenario in valuation_analyzer.scenario_results:
                    scenario_data = valuation_analyzer.scenario_results[scenario]
                    value = scenario_data.get('value_per_share', 0)
                    if value > 0:
                        self.audit_results['financial_checks'].append(
                            f"✓ {scenario.capitalize()} case calculated: {value:.2f}"
                        )
                    else:
                        error = f"{scenario.capitalize()} case has invalid value"
                        self.audit_results['errors'].append(error)
                        print(f"  ✗ ERROR: {error}")
        
        # Check sensitivity analysis
        if valuation_analyzer.sensitivity_results:
            for key, result_df in valuation_analyzer.sensitivity_results.items():
                if isinstance(result_df, pd.DataFrame) and not result_df.empty:
                    self.audit_results['financial_checks'].append(
                        f"✓ {key.capitalize()} sensitivity analysis completed"
                    )
    
    def _audit_excel_file(self, excel_file_path: str):
        """Audit Excel file for formula correctness and consistency"""
        print("\n[Excel Audit] Checking Excel file...")
        
        try:
            wb = load_workbook(excel_file_path, data_only=False)
            
            # Check for required sheets
            required_sheets = [
                'Executive Summary', 'Historical Financials', 'DCF Assumptions',
                'Income Statement Projections', 'FCFF Calculation', 'WACC Calculation',
                'Terminal Value', 'DCF Valuation'
            ]
            
            existing_sheets = wb.sheetnames
            for sheet_name in required_sheets:
                if sheet_name in existing_sheets:
                    self.audit_results['excel_checks'].append(f"✓ Sheet '{sheet_name}' exists")
                else:
                    warning = f"Sheet '{sheet_name}' not found"
                    self.audit_results['warnings'].append(warning)
                    print(f"  ⚠ WARNING: {warning}")
            
            # Check for formulas in key sheets
            key_sheets = ['FCFF Calculation', 'WACC Calculation', 'DCF Valuation']
            for sheet_name in key_sheets:
                if sheet_name in existing_sheets:
                    ws = wb[sheet_name]
                    formula_count = 0
                    for row in ws.iter_rows():
                        for cell in row:
                            if cell.value and str(cell.value).startswith('='):
                                formula_count += 1
                    
                    if formula_count > 0:
                        self.audit_results['excel_checks'].append(
                            f"✓ '{sheet_name}' contains {formula_count} formulas"
                        )
                    else:
                        warning = f"'{sheet_name}' contains no formulas"
                        self.audit_results['warnings'].append(warning)
                        print(f"  ⚠ WARNING: {warning}")
            
            wb.close()
            
        except Exception as e:
            error = f"Error auditing Excel file: {str(e)}"
            self.audit_results['errors'].append(error)
            print(f"  ✗ ERROR: {error}")
    
    def _audit_code_quality(self):
        """Audit code quality and technical soundness"""
        print("\n[Technical Audit] Checking code quality...")
        
        # Check for required modules
        required_modules = [
            'data_collection', 'financial_analysis', 'dcf_model',
            'valuation_analysis', 'excel_generator', 'audit_system'
        ]
        
        for module_name in required_modules:
            module_path = f"{module_name}.py"
            if os.path.exists(module_path):
                self.audit_results['technical_checks'].append(f"✓ Module '{module_name}.py' exists")
            else:
                error = f"Module '{module_name}.py' not found"
                self.audit_results['errors'].append(error)
                print(f"  ✗ ERROR: {error}")
        
        # Check config file
        if os.path.exists('config.py'):
            self.audit_results['technical_checks'].append("✓ Config file exists")
        else:
            error = "Config file not found"
            self.audit_results['errors'].append(error)
            print(f"  ✗ ERROR: {error}")
    
    def _print_audit_summary(self):
        """Print audit summary"""
        print("\n" + "=" * 60)
        print("AUDIT SUMMARY")
        print("=" * 60)
        
        total_checks = (len(self.audit_results['financial_checks']) +
                       len(self.audit_results['technical_checks']) +
                       len(self.audit_results['excel_checks']))
        
        print(f"\n✓ Passed Checks: {total_checks}")
        print(f"⚠ Warnings: {len(self.audit_results['warnings'])}")
        print(f"✗ Errors: {len(self.audit_results['errors'])}")
        
        if self.audit_results['warnings']:
            print("\nWarnings:")
            for warning in self.audit_results['warnings']:
                print(f"  - {warning}")
        
        if self.audit_results['errors']:
            print("\nErrors:")
            for error in self.audit_results['errors']:
                print(f"  - {error}")
        
        if len(self.audit_results['errors']) == 0:
            print("\n✓ Audit passed! No errors found.")
        else:
            print(f"\n✗ Audit failed with {len(self.audit_results['errors'])} error(s).")
        
        print("=" * 60)
    
    def get_audit_status(self) -> Tuple[bool, str]:
        """
        Get audit status
        
        Returns:
            Tuple of (is_passed, status_message)
        """
        error_count = len(self.audit_results['errors'])
        warning_count = len(self.audit_results['warnings'])
        
        if error_count == 0:
            if warning_count == 0:
                return True, "All checks passed"
            else:
                return True, f"Passed with {warning_count} warning(s)"
        else:
            return False, f"Failed with {error_count} error(s) and {warning_count} warning(s)"

