"""
Excel Output Generator
Creates professionally formatted Excel workbook with all DCF model components,
formulas, and visualizations
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from typing import Dict, Optional
import os
import sys

# Add parent directory to path for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config
from utils.formatting import (
    get_header_style, get_input_style, get_calculation_style,
    get_formula_style, apply_style_to_cell, format_number, auto_adjust_column_width
)


class ExcelGenerator:
    """Generates comprehensive Excel output for DCF model"""
    
    def __init__(self, financial_analyzer, dcf_model, valuation_analyzer, 
                 data_collector, final_recommendation: Dict):
        """
        Initialize Excel generator
        
        Args:
            financial_analyzer: FinancialAnalyzer instance
            dcf_model: DCFModel instance
            valuation_analyzer: ValuationAnalyzer instance
            data_collector: DataCollector instance
            final_recommendation: Final recommendation dictionary
        """
        self.financial_analyzer = financial_analyzer
        self.dcf_model = dcf_model
        self.valuation_analyzer = valuation_analyzer
        self.data_collector = data_collector
        self.final_recommendation = final_recommendation
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet
        
    def generate_excel(self, output_path: str = None) -> str:
        """
        Generate complete Excel workbook
        
        Args:
            output_path: Output file path (defaults to config)
        
        Returns:
            Path to generated Excel file
        """
        if output_path is None:
            os.makedirs(config.OUTPUT_DIR, exist_ok=True)
            output_path = config.EXCEL_OUTPUT_FILE
        
        print(f"Generating Excel workbook: {output_path}")
        
        # Validate data before generation
        validation_errors = self._validate_data()
        if validation_errors:
            print("  Warnings during data validation:")
            for error in validation_errors:
                print(f"    - {error}")
        
        # Create all sheets
        self._create_executive_summary()
        self._create_data_sources()
        self._create_historical_financials()
        self._create_financial_analysis()
        self._create_dcf_assumptions()
        self._create_revenue_model()
        self._create_income_statement_projections()
        self._create_balance_sheet_projections()
        self._create_cash_flow_projections()
        self._create_fcff_calculation()
        self._create_wacc_calculation()
        self._create_terminal_value()
        self._create_dcf_valuation()
        self._create_sensitivity_analysis()
        self._create_scenario_analysis()
        self._create_relative_valuation()
        self._create_summary()
        
        # Save workbook
        self.wb.save(output_path)
        print(f"Excel workbook saved: {output_path}")
        
        return output_path
    
    def _validate_data(self) -> list:
        """
        Validate that required data exists before Excel generation
        
        Returns:
            List of validation error messages (empty if all valid)
        """
        errors = []
        
        # Check financial statements
        if self.financial_analyzer.normalized_income_stmt.empty:
            errors.append("Income statement is empty - historical data may not be available")
        elif 'Revenue' not in self.financial_analyzer.normalized_income_stmt.columns:
            errors.append("Revenue column not found in income statement")
        
        if self.financial_analyzer.normalized_balance_sheet.empty:
            errors.append("Balance sheet is empty - historical data may not be available")
        
        if self.financial_analyzer.normalized_cash_flow.empty:
            errors.append("Cash flow statement is empty - historical data may not be available")
        
        # Check DCF projections
        if self.dcf_model.revenue_projections is None or len(self.dcf_model.revenue_projections) == 0:
            errors.append("Revenue projections are missing")
        
        if self.dcf_model.income_projections is None or self.dcf_model.income_projections.empty:
            errors.append("Income statement projections are missing")
        
        if self.dcf_model.wacc is None:
            errors.append("WACC has not been calculated")
        
        # Check final recommendation
        if not self.final_recommendation:
            errors.append("Final recommendation data is missing")
        
        return errors
    
    def _create_executive_summary(self):
        """Create Executive Summary sheet"""
        ws = self.wb.create_sheet("Executive Summary", 0)
        
        row = 1
        
        # Title
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = f"{config.COMPANY_NAME} ({config.COMPANY_TICKER}) - DCF Valuation"
        cell.font = Font(size=16, bold=True)
        row += 2
        
        # Key Metrics
        ws[f'A{row}'] = "Investment Recommendation"
        ws[f'B{row}'] = self.final_recommendation.get('recommendation', 'N/A')
        apply_style_to_cell(ws[f'A{row}'], get_header_style())
        apply_style_to_cell(ws[f'B{row}'], get_input_style())
        row += 1
        
        ws[f'A{row}'] = "Current Price"
        ws[f'B{row}'] = self.final_recommendation.get('current_price', 0)
        apply_style_to_cell(ws[f'A{row}'], get_header_style())
        apply_style_to_cell(ws[f'B{row}'], get_input_style())
        ws[f'B{row}'].number_format = config.EXCEL_FORMATTING['currency_format']
        row += 1
        
        ws[f'A{row}'] = "Target Price"
        ws[f'B{row}'] = self.final_recommendation.get('target_price', 0)
        apply_style_to_cell(ws[f'A{row}'], get_header_style())
        apply_style_to_cell(ws[f'B{row}'], get_input_style())
        ws[f'B{row}'].number_format = config.EXCEL_FORMATTING['currency_format']
        row += 1
        
        ws[f'A{row}'] = "Upside/Downside"
        upside = self.final_recommendation.get('upside_downside_pct', 0)
        ws[f'B{row}'] = upside / 100
        apply_style_to_cell(ws[f'A{row}'], get_header_style())
        apply_style_to_cell(ws[f'B{row}'], get_input_style())
        ws[f'B{row}'].number_format = config.EXCEL_FORMATTING['percentage_format']
        row += 2
        
        # DCF Value
        ws[f'A{row}'] = "DCF Value per Share"
        ws[f'B{row}'] = self.final_recommendation.get('dcf_value', 0)
        apply_style_to_cell(ws[f'A{row}'], get_header_style())
        apply_style_to_cell(ws[f'B{row}'], get_calculation_style())
        ws[f'B{row}'].number_format = config.EXCEL_FORMATTING['currency_format']
        row += 1
        
        # Relative Value
        if self.final_recommendation.get('relative_value'):
            ws[f'A{row}'] = "Relative Value per Share"
            ws[f'B{row}'] = self.final_recommendation['relative_value']
            apply_style_to_cell(ws[f'A{row}'], get_header_style())
            apply_style_to_cell(ws[f'B{row}'], get_calculation_style())
            ws[f'B{row}'].number_format = config.EXCEL_FORMATTING['currency_format']
            row += 1
        
        # Key Assumptions
        row += 1
        ws[f'A{row}'] = "Key Assumptions"
        apply_style_to_cell(ws[f'A{row}'], get_header_style())
        row += 1
        
        ws[f'A{row}'] = "WACC"
        ws[f'B{row}'] = self.dcf_model.wacc if self.dcf_model.wacc else 0
        apply_style_to_cell(ws[f'A{row}'], get_header_style())
        apply_style_to_cell(ws[f'B{row}'], get_input_style())
        ws[f'B{row}'].number_format = config.EXCEL_FORMATTING['percentage_format']
        row += 1
        
        ws[f'A{row}'] = "Terminal Growth Rate"
        terminal_growth = config.DEFAULT_ASSUMPTIONS['terminal_growth_rate']
        ws[f'B{row}'] = terminal_growth
        apply_style_to_cell(ws[f'A{row}'], get_header_style())
        apply_style_to_cell(ws[f'B{row}'], get_input_style())
        ws[f'B{row}'].number_format = config.EXCEL_FORMATTING['percentage_format']
        
        auto_adjust_column_width(ws)
    
    def _create_data_sources(self):
        """Create Data Sources sheet"""
        ws = self.wb.create_sheet("Data Sources")
        
        row = 1
        ws[f'A{row}'] = "Data Source"
        ws[f'B{row}'] = "Description"
        ws[f'C{row}'] = "Date Collected"
        for col in ['A', 'B', 'C']:
            apply_style_to_cell(ws[f'{col}{row}'], get_header_style())
        row += 1
        
        sources = [
            ("Yahoo Finance", "Financial statements, market data, beta", "Auto"),
            ("Company IR Website", "Annual reports, investor presentations", "Manual"),
            ("Market Data", "Current price, market cap, shares outstanding", "Auto"),
            ("Macro Data", "Risk-free rate, equity risk premium", "Manual"),
        ]
        
        for source, desc, date in sources:
            ws[f'A{row}'] = source
            ws[f'B{row}'] = desc
            ws[f'C{row}'] = date
            for col in ['A', 'B', 'C']:
                apply_style_to_cell(ws[f'{col}{row}'], get_calculation_style())
            row += 1
        
        auto_adjust_column_width(ws)
    
    def _create_historical_financials(self):
        """Create Historical Financials sheet"""
        ws = self.wb.create_sheet("Historical Financials")
        
        # Income Statement
        row = 1
        ws[f'A{row}'] = "Income Statement (Historical)"
        apply_style_to_cell(ws[f'A{row}'], get_header_style())
        row += 2
        
        if not self.financial_analyzer.normalized_income_stmt.empty:
            income_stmt = self.financial_analyzer.normalized_income_stmt
            
            # Structure: rows are periods (dates), columns are line items
            # Headers: Date column + line items
            ws[f'A{row}'] = "Period"
            col = 2
            
            # Get all date periods (rows in the DataFrame)
            periods = income_stmt['Date'].tolist() if 'Date' in income_stmt.columns else []
            
            # Write line item headers
            line_items = ['Revenue', 'Cost of Revenue', 'Gross Profit', 
                         'EBIT', 'EBITDA', 'Net Income']
            available_line_items = [item for item in line_items if item in income_stmt.columns]
            
            for line_item in available_line_items:
                ws.cell(row, col, line_item)
                apply_style_to_cell(ws.cell(row, col), get_header_style())
                col += 1
            row += 1
            
            # Write data: each row is a period, each column is a line item
            for period_idx, period_date in enumerate(periods):
                # Write period/date
                try:
                    if hasattr(period_date, 'strftime'):
                        period_str = period_date.strftime('%Y-%m-%d')
                    else:
                        period_str = str(period_date)
                except:
                    period_str = str(period_date)
                ws.cell(row, 1, period_str)
                apply_style_to_cell(ws.cell(row, 1), get_header_style())
                
                # Write line item values for this period
                col = 2
                for line_item in available_line_items:
                    if line_item in income_stmt.columns:
                        value = income_stmt[line_item].iloc[period_idx]
                        ws.cell(row, col, value)
                        apply_style_to_cell(ws.cell(row, col), get_calculation_style())
                        ws.cell(row, col).number_format = config.EXCEL_FORMATTING['number_format']
                    col += 1
                row += 1
        
        # Balance Sheet Section
        row += 2
        ws[f'A{row}'] = "Balance Sheet (Historical)"
        apply_style_to_cell(ws[f'A{row}'], get_header_style())
        row += 2
        
        if not self.financial_analyzer.normalized_balance_sheet.empty:
            balance_sheet = self.financial_analyzer.normalized_balance_sheet
            
            ws[f'A{row}'] = "Period"
            col = 2
            
            periods = balance_sheet['Date'].tolist() if 'Date' in balance_sheet.columns else []
            line_items = ['Total Assets', 'Total Liabilities', 'Total Equity', 
                         'Current Assets', 'Current Liabilities', 'Total Debt',
                         'Cash and Cash Equivalents']
            available_line_items = [item for item in line_items if item in balance_sheet.columns]
            
            for line_item in available_line_items:
                ws.cell(row, col, line_item)
                apply_style_to_cell(ws.cell(row, col), get_header_style())
                col += 1
            row += 1
            
            for period_idx, period_date in enumerate(periods):
                if isinstance(period_date, pd.Timestamp):
                    period_str = period_date.strftime('%Y-%m-%d')
                else:
                    period_str = str(period_date)
                ws.cell(row, 1, period_str)
                apply_style_to_cell(ws.cell(row, 1), get_header_style())
                
                col = 2
                for line_item in available_line_items:
                    if line_item in balance_sheet.columns:
                        value = balance_sheet[line_item].iloc[period_idx]
                        ws.cell(row, col, value)
                        apply_style_to_cell(ws.cell(row, col), get_calculation_style())
                        ws.cell(row, col).number_format = config.EXCEL_FORMATTING['number_format']
                    col += 1
                row += 1
        
        # Cash Flow Section
        row += 2
        ws[f'A{row}'] = "Cash Flow Statement (Historical)"
        apply_style_to_cell(ws[f'A{row}'], get_header_style())
        row += 2
        
        if not self.financial_analyzer.normalized_cash_flow.empty:
            cash_flow = self.financial_analyzer.normalized_cash_flow
            
            ws[f'A{row}'] = "Period"
            col = 2
            
            periods = cash_flow['Date'].tolist() if 'Date' in cash_flow.columns else []
            line_items = ['Operating Cash Flow', 'Investing Cash Flow', 
                         'Financing Cash Flow', 'Capital Expenditures', 
                         'Net Change in Cash']
            available_line_items = [item for item in line_items if item in cash_flow.columns]
            
            for line_item in available_line_items:
                ws.cell(row, col, line_item)
                apply_style_to_cell(ws.cell(row, col), get_header_style())
                col += 1
            row += 1
            
            for period_idx, period_date in enumerate(periods):
                if isinstance(period_date, pd.Timestamp):
                    period_str = period_date.strftime('%Y-%m-%d')
                else:
                    period_str = str(period_date)
                ws.cell(row, 1, period_str)
                apply_style_to_cell(ws.cell(row, 1), get_header_style())
                
                col = 2
                for line_item in available_line_items:
                    if line_item in cash_flow.columns:
                        value = cash_flow[line_item].iloc[period_idx]
                        ws.cell(row, col, value)
                        apply_style_to_cell(ws.cell(row, col), get_calculation_style())
                        ws.cell(row, col).number_format = config.EXCEL_FORMATTING['number_format']
                    col += 1
                row += 1
        
        auto_adjust_column_width(ws)
    
    def _create_financial_analysis(self):
        """Create Financial Analysis sheet"""
        ws = self.wb.create_sheet("Financial Analysis")
        
        row = 1
        ws[f'A{row}'] = "Financial Ratios"
        apply_style_to_cell(ws[f'A{row}'], get_header_style())
        row += 2
        
        # Margins
        if 'gross_margin' in self.financial_analyzer.ratios:
            ws[f'A{row}'] = "Gross Margin"
            margins = self.financial_analyzer.ratios['gross_margin']
            if margins:
                ws[f'B{row}'] = margins[-1] if isinstance(margins, list) else margins
                apply_style_to_cell(ws[f'A{row}'], get_header_style())
                apply_style_to_cell(ws[f'B{row}'], get_calculation_style())
                ws[f'B{row}'].number_format = config.EXCEL_FORMATTING['percentage_format']
                row += 1
        
        if 'ebit_margin' in self.financial_analyzer.ratios:
            ws[f'A{row}'] = "EBIT Margin"
            margins = self.financial_analyzer.ratios['ebit_margin']
            if margins:
                ws[f'B{row}'] = margins[-1] if isinstance(margins, list) else margins
                apply_style_to_cell(ws[f'A{row}'], get_header_style())
                apply_style_to_cell(ws[f'B{row}'], get_calculation_style())
                ws[f'B{row}'].number_format = config.EXCEL_FORMATTING['percentage_format']
                row += 1
        
        auto_adjust_column_width(ws)
    
    def _create_dcf_assumptions(self):
        """Create DCF Assumptions sheet"""
        ws = self.wb.create_sheet("DCF Assumptions")
        
        row = 1
        ws[f'A{row}'] = "DCF Model Assumptions"
        apply_style_to_cell(ws[f'A{row}'], get_header_style())
        row += 2
        
        assumptions = [
            ("Forecast Period (Years)", self.dcf_model.forecast_years),
            ("Terminal Growth Rate", config.DEFAULT_ASSUMPTIONS['terminal_growth_rate']),
            ("Risk-Free Rate", self.dcf_model.macro_data.get('risk_free_rate', 0.025)),
            ("Equity Risk Premium", self.dcf_model.macro_data.get('equity_risk_premium', 0.05)),
            ("Beta", self.dcf_model.market_data.get('beta', 1.0)),
        ]
        
        for label, value in assumptions:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            apply_style_to_cell(ws[f'A{row}'], get_header_style())
            apply_style_to_cell(ws[f'B{row}'], get_input_style())
            if 'Rate' in label or 'Premium' in label or 'Growth' in label:
                ws[f'B{row}'].number_format = config.EXCEL_FORMATTING['percentage_format']
            row += 1
        
        auto_adjust_column_width(ws)
    
    def _create_revenue_model(self):
        """Create Revenue Model sheet"""
        ws = self.wb.create_sheet("Revenue Model")
        
        row = 1
        ws[f'A{row}'] = "Revenue Projections"
        apply_style_to_cell(ws[f'A{row}'], get_header_style())
        row += 2
        
        if self.dcf_model.revenue_projections is not None:
            ws[f'A{row}'] = "Year"
            ws[f'B{row}'] = "Revenue"
            ws[f'C{row}'] = "Growth Rate"
            for col in ['A', 'B', 'C']:
                apply_style_to_cell(ws[f'{col}{row}'], get_header_style())
            row += 1
            
            for year, revenue in self.dcf_model.revenue_projections.items():
                ws[f'A{row}'] = year
                ws[f'B{row}'] = revenue
                if row > 2:  # Calculate growth
                    prev_revenue = ws[f'B{row-1}'].value
                    if prev_revenue:
                        ws[f'C{row}'] = f"=IF(B{row-1}<>0, (B{row}-B{row-1})/B{row-1}, 0)"
                        ws[f'C{row}'].number_format = config.EXCEL_FORMATTING['percentage_format']
                apply_style_to_cell(ws[f'A{row}'], get_header_style())
                apply_style_to_cell(ws[f'B{row}'], get_calculation_style())
                apply_style_to_cell(ws[f'C{row}'], get_formula_style())
                ws[f'B{row}'].number_format = config.EXCEL_FORMATTING['number_format']
                row += 1
        
        auto_adjust_column_width(ws)
    
    def _create_income_statement_projections(self):
        """Create Income Statement Projections sheet"""
        ws = self.wb.create_sheet("Income Statement Projections")
        
        row = 1
        ws[f'A{row}'] = "Income Statement - 5 Year Forecast"
        apply_style_to_cell(ws[f'A{row}'], get_header_style())
        row += 2
        
        if self.dcf_model.income_projections is not None:
            income = self.dcf_model.income_projections
            
            # Headers
            ws[f'A{row}'] = "Line Item"
            col = 2
            for year in income.index:
                ws.cell(row, col, f"Year {year}")
                apply_style_to_cell(ws.cell(row, col), get_header_style())
                col += 1
            row += 1
            
            # Data rows
            line_items = ['Revenue', 'Cost of Revenue', 'Gross Profit', 
                         'Operating Expenses', 'EBIT', 'Depreciation', 'EBITDA',
                         'Interest Expense', 'Income Before Tax', 'Income Tax Expense', 'Net Income']
            
            for line_item in line_items:
                if line_item in income.columns:
                    ws[f'A{row}'] = line_item
                    apply_style_to_cell(ws[f'A{row}'], get_header_style())
                    col = 2
                    for year in income.index:
                        value = income.loc[year, line_item]
                        ws.cell(row, col, value)
                        apply_style_to_cell(ws.cell(row, col), get_calculation_style())
                        ws.cell(row, col).number_format = config.EXCEL_FORMATTING['number_format']
                        col += 1
                    row += 1
        
        auto_adjust_column_width(ws)
    
    def _create_balance_sheet_projections(self):
        """Create Balance Sheet Projections sheet"""
        ws = self.wb.create_sheet("Balance Sheet Projections")
        
        row = 1
        ws[f'A{row}'] = "Balance Sheet - 5 Year Forecast"
        apply_style_to_cell(ws[f'A{row}'], get_header_style())
        row += 2
        
        if self.dcf_model.balance_sheet_projections is not None:
            balance = self.dcf_model.balance_sheet_projections
            
            ws[f'A{row}'] = "Line Item"
            col = 2
            for year in balance.index:
                ws.cell(row, col, f"Year {year}")
                apply_style_to_cell(ws.cell(row, col), get_header_style())
                col += 1
            row += 1
            
            for line_item in balance.columns:
                ws[f'A{row}'] = line_item
                apply_style_to_cell(ws[f'A{row}'], get_header_style())
                col = 2
                for year in balance.index:
                    value = balance.loc[year, line_item]
                    ws.cell(row, col, value)
                    apply_style_to_cell(ws.cell(row, col), get_calculation_style())
                    ws.cell(row, col).number_format = config.EXCEL_FORMATTING['number_format']
                    col += 1
                row += 1
        
        auto_adjust_column_width(ws)
    
    def _create_cash_flow_projections(self):
        """Create Cash Flow Projections sheet"""
        ws = self.wb.create_sheet("Cash Flow Projections")
        
        row = 1
        ws[f'A{row}'] = "Cash Flow Statement - 5 Year Forecast"
        apply_style_to_cell(ws[f'A{row}'], get_header_style())
        row += 2
        
        if self.dcf_model.cash_flow_projections is not None:
            cash_flow = self.dcf_model.cash_flow_projections
            
            ws[f'A{row}'] = "Line Item"
            col = 2
            for year in cash_flow.index:
                ws.cell(row, col, f"Year {year}")
                apply_style_to_cell(ws.cell(row, col), get_header_style())
                col += 1
            row += 1
            
            for line_item in cash_flow.columns:
                ws[f'A{row}'] = line_item
                apply_style_to_cell(ws[f'A{row}'], get_header_style())
                col = 2
                for year in cash_flow.index:
                    value = cash_flow.loc[year, line_item]
                    ws.cell(row, col, value)
                    apply_style_to_cell(ws.cell(row, col), get_calculation_style())
                    ws.cell(row, col).number_format = config.EXCEL_FORMATTING['number_format']
                    col += 1
                row += 1
        
        auto_adjust_column_width(ws)
    
    def _create_fcff_calculation(self):
        """Create FCFF Calculation sheet with formulas"""
        ws = self.wb.create_sheet("FCFF Calculation")
        
        row = 1
        ws[f'A{row}'] = "Free Cash Flow to Firm (FCFF) Calculation"
        apply_style_to_cell(ws[f'A{row}'], get_header_style())
        row += 2
        
        ws[f'A{row}'] = "Year"
        ws[f'B{row}'] = "EBIT"
        ws[f'C{row}'] = "EBIT(1-t)"
        ws[f'D{row}'] = "Depreciation"
        ws[f'E{row}'] = "CapEx"
        ws[f'F{row}'] = "ΔNWC"
        ws[f'G{row}'] = "FCFF"
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            apply_style_to_cell(ws[f'{col}{row}'], get_header_style())
        row += 1
        
        if self.dcf_model.fcff_projections is not None:
            tax_rate = 0.25  # Should get from assumptions
            income = self.dcf_model.income_projections
            cash_flow = self.dcf_model.cash_flow_projections
            balance = self.dcf_model.balance_sheet_projections
            
            for i, year in enumerate(income.index, start=1):
                ws[f'A{row}'] = f"Year {year}"
                
                # EBIT
                ebit_cell = f"='Income Statement Projections'!{get_column_letter(income.columns.get_loc('EBIT')+2)}{income.index.get_loc(year)+3}"
                ws[f'B{row}'] = f"={ebit_cell}"
                
                # EBIT(1-t)
                ws[f'C{row}'] = f"=B{row}*(1-{tax_rate})"
                
                # Depreciation
                dep_cell = f"='Income Statement Projections'!{get_column_letter(income.columns.get_loc('Depreciation')+2)}{income.index.get_loc(year)+3}"
                ws[f'D{row}'] = f"={dep_cell}"
                
                # CapEx (negative)
                capex_cell = f"='Cash Flow Projections'!{get_column_letter(cash_flow.columns.get_loc('Capital Expenditures')+2)}{cash_flow.index.get_loc(year)+3}"
                ws[f'E{row}'] = f"={capex_cell}"
                
                # ΔNWC
                wc_cell = f"='Balance Sheet Projections'!{get_column_letter(balance.columns.get_loc('Change in WC')+2)}{balance.index.get_loc(year)+3}"
                ws[f'F{row}'] = f"={wc_cell}"
                
                # FCFF = EBIT(1-t) + Depreciation - CapEx - ΔNWC
                ws[f'G{row}'] = f"=C{row}+D{row}+E{row}-F{row}"
                
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                    if col == 'A':
                        apply_style_to_cell(ws[f'{col}{row}'], get_header_style())
                    else:
                        apply_style_to_cell(ws[f'{col}{row}'], get_formula_style())
                        ws[f'{col}{row}'].number_format = config.EXCEL_FORMATTING['number_format']
                row += 1
        
        auto_adjust_column_width(ws)
    
    def _create_wacc_calculation(self):
        """Create WACC Calculation sheet"""
        ws = self.wb.create_sheet("WACC Calculation")
        
        row = 1
        ws[f'A{row}'] = "Weighted Average Cost of Capital (WACC) Calculation"
        apply_style_to_cell(ws[f'A{row}'], get_header_style())
        row += 2
        
        # Cost of Equity (CAPM)
        ws[f'A{row}'] = "Cost of Equity (CAPM)"
        apply_style_to_cell(ws[f'A{row}'], get_header_style())
        row += 1
        
        ws[f'B{row}'] = "Risk-Free Rate"
        ws[f'C{row}'] = self.dcf_model.macro_data.get('risk_free_rate', 0.025)
        apply_style_to_cell(ws[f'B{row}'], get_header_style())
        apply_style_to_cell(ws[f'C{row}'], get_input_style())
        ws[f'C{row}'].number_format = config.EXCEL_FORMATTING['percentage_format']
        row += 1
        
        ws[f'B{row}'] = "Beta"
        ws[f'C{row}'] = self.dcf_model.market_data.get('beta', 1.0)
        apply_style_to_cell(ws[f'B{row}'], get_header_style())
        apply_style_to_cell(ws[f'C{row}'], get_input_style())
        row += 1
        
        ws[f'B{row}'] = "Equity Risk Premium"
        ws[f'C{row}'] = self.dcf_model.macro_data.get('equity_risk_premium', 0.05)
        apply_style_to_cell(ws[f'B{row}'], get_header_style())
        apply_style_to_cell(ws[f'C{row}'], get_input_style())
        ws[f'C{row}'].number_format = config.EXCEL_FORMATTING['percentage_format']
        row += 1
        
        ws[f'B{row}'] = "Cost of Equity"
        ws[f'C{row}'] = f"=C{row-3}+C{row-2}*C{row-1}"
        apply_style_to_cell(ws[f'B{row}'], get_header_style())
        apply_style_to_cell(ws[f'C{row}'], get_formula_style())
        ws[f'C{row}'].number_format = config.EXCEL_FORMATTING['percentage_format']
        cost_equity_row = row
        row += 2
        
        # Cost of Debt
        ws[f'A{row}'] = "Cost of Debt"
        apply_style_to_cell(ws[f'A{row}'], get_header_style())
        row += 1
        
        ws[f'B{row}'] = "Cost of Debt"
        ws[f'C{row}'] = self.dcf_model._calculate_cost_of_debt()
        apply_style_to_cell(ws[f'B{row}'], get_header_style())
        apply_style_to_cell(ws[f'C{row}'], get_input_style())
        ws[f'C{row}'].number_format = config.EXCEL_FORMATTING['percentage_format']
        cost_debt_row = row
        row += 2
        
        # Capital Structure
        ws[f'A{row}'] = "Capital Structure"
        apply_style_to_cell(ws[f'A{row}'], get_header_style())
        row += 1
        
        equity_weight, debt_weight = self.dcf_model._get_capital_structure()
        ws[f'B{row}'] = "Equity Weight"
        ws[f'C{row}'] = equity_weight
        apply_style_to_cell(ws[f'B{row}'], get_header_style())
        apply_style_to_cell(ws[f'C{row}'], get_input_style())
        ws[f'C{row}'].number_format = config.EXCEL_FORMATTING['percentage_format']
        equity_weight_row = row
        row += 1
        
        ws[f'B{row}'] = "Debt Weight"
        ws[f'C{row}'] = debt_weight
        apply_style_to_cell(ws[f'B{row}'], get_header_style())
        apply_style_to_cell(ws[f'C{row}'], get_input_style())
        ws[f'C{row}'].number_format = config.EXCEL_FORMATTING['percentage_format']
        debt_weight_row = row
        row += 2
        
        # Tax Rate
        ws[f'B{row}'] = "Tax Rate"
        tax_rate = 0.25
        ws[f'C{row}'] = tax_rate
        apply_style_to_cell(ws[f'B{row}'], get_header_style())
        apply_style_to_cell(ws[f'C{row}'], get_input_style())
        ws[f'C{row}'].number_format = config.EXCEL_FORMATTING['percentage_format']
        tax_rate_row = row
        row += 2
        
        # WACC
        ws[f'A{row}'] = "WACC"
        ws[f'B{row}'] = f"=C{equity_weight_row}*C{cost_equity_row}+C{debt_weight_row}*C{cost_debt_row}*(1-C{tax_rate_row})"
        apply_style_to_cell(ws[f'A{row}'], get_header_style())
        apply_style_to_cell(ws[f'B{row}'], get_formula_style())
        ws[f'B{row}'].number_format = config.EXCEL_FORMATTING['percentage_format']
        
        auto_adjust_column_width(ws)
    
    def _create_terminal_value(self):
        """Create Terminal Value sheet"""
        ws = self.wb.create_sheet("Terminal Value")
        
        row = 1
        ws[f'A{row}'] = "Terminal Value Calculation"
        apply_style_to_cell(ws[f'A{row}'], get_header_style())
        row += 2
        
        # Perpetuity Growth Method
        ws[f'A{row}'] = "Perpetuity Growth Method"
        apply_style_to_cell(ws[f'A{row}'], get_header_style())
        row += 1
        
        final_fcff = self.dcf_model.fcff_projections.iloc[-1] if self.dcf_model.fcff_projections is not None else 0
        terminal_growth = config.DEFAULT_ASSUMPTIONS['terminal_growth_rate']
        wacc = self.dcf_model.wacc if self.dcf_model.wacc else 0.10
        
        ws[f'B{row}'] = "Final Year FCFF"
        ws[f'C{row}'] = final_fcff
        apply_style_to_cell(ws[f'B{row}'], get_header_style())
        apply_style_to_cell(ws[f'C{row}'], get_calculation_style())
        ws[f'C{row}'].number_format = config.EXCEL_FORMATTING['number_format']
        row += 1
        
        ws[f'B{row}'] = "Terminal Growth Rate"
        ws[f'C{row}'] = terminal_growth
        apply_style_to_cell(ws[f'B{row}'], get_header_style())
        apply_style_to_cell(ws[f'C{row}'], get_input_style())
        ws[f'C{row}'].number_format = config.EXCEL_FORMATTING['percentage_format']
        row += 1
        
        ws[f'B{row}'] = "WACC"
        ws[f'C{row}'] = wacc
        apply_style_to_cell(ws[f'B{row}'], get_header_style())
        apply_style_to_cell(ws[f'C{row}'], get_input_style())
        ws[f'C{row}'].number_format = config.EXCEL_FORMATTING['percentage_format']
        row += 1
        
        ws[f'B{row}'] = "Terminal Value (Perpetuity)"
        ws[f'C{row}'] = f"=C{row-3}*(1+C{row-2})/(C{row-1}-C{row-2})"
        apply_style_to_cell(ws[f'B{row}'], get_header_style())
        apply_style_to_cell(ws[f'C{row}'], get_formula_style())
        ws[f'C{row}'].number_format = config.EXCEL_FORMATTING['number_format']
        
        auto_adjust_column_width(ws)
    
    def _create_dcf_valuation(self):
        """Create DCF Valuation sheet"""
        ws = self.wb.create_sheet("DCF Valuation")
        
        row = 1
        ws[f'A{row}'] = "DCF Valuation Summary"
        apply_style_to_cell(ws[f'A{row}'], get_header_style())
        row += 2
        
        # PV of FCFF
        ws[f'A{row}'] = "Present Value of FCFF"
        pv_fcff = self.dcf_model.enterprise_value - (self.dcf_model.terminal_value / ((1 + self.dcf_model.wacc) ** (self.dcf_model.forecast_years - 0.5))) if self.dcf_model.enterprise_value else 0
        ws[f'B{row}'] = pv_fcff
        apply_style_to_cell(ws[f'A{row}'], get_header_style())
        apply_style_to_cell(ws[f'B{row}'], get_calculation_style())
        ws[f'B{row}'].number_format = config.EXCEL_FORMATTING['number_format']
        row += 1
        
        # PV of Terminal Value
        ws[f'A{row}'] = "Present Value of Terminal Value"
        pv_tv = self.dcf_model.terminal_value / ((1 + self.dcf_model.wacc) ** (self.dcf_model.forecast_years - 0.5)) if self.dcf_model.terminal_value else 0
        ws[f'B{row}'] = pv_tv
        apply_style_to_cell(ws[f'A{row}'], get_header_style())
        apply_style_to_cell(ws[f'B{row}'], get_calculation_style())
        ws[f'B{row}'].number_format = config.EXCEL_FORMATTING['number_format']
        row += 1
        
        # Enterprise Value
        ws[f'A{row}'] = "Enterprise Value"
        ws[f'B{row}'] = f"=B{row-2}+B{row-1}"
        apply_style_to_cell(ws[f'A{row}'], get_header_style())
        apply_style_to_cell(ws[f'B{row}'], get_formula_style())
        ws[f'B{row}'].number_format = config.EXCEL_FORMATTING['number_format']
        row += 1
        
        # Net Debt
        net_debt = 0
        if not self.financial_analyzer.normalized_balance_sheet.empty:
            latest = self.financial_analyzer.normalized_balance_sheet.iloc[-1]
            if 'Net Debt' in latest:
                net_debt = latest['Net Debt']
        
        ws[f'A{row}'] = "Net Debt"
        ws[f'B{row}'] = net_debt
        apply_style_to_cell(ws[f'A{row}'], get_header_style())
        apply_style_to_cell(ws[f'B{row}'], get_calculation_style())
        ws[f'B{row}'].number_format = config.EXCEL_FORMATTING['number_format']
        row += 1
        
        # Equity Value
        ws[f'A{row}'] = "Equity Value"
        ws[f'B{row}'] = f"=B{row-2}-B{row-1}"
        apply_style_to_cell(ws[f'A{row}'], get_header_style())
        apply_style_to_cell(ws[f'B{row}'], get_formula_style())
        ws[f'B{row}'].number_format = config.EXCEL_FORMATTING['number_format']
        row += 1
        
        # Shares Outstanding
        shares = self.dcf_model.market_data.get('shares_outstanding', 1)
        ws[f'A{row}'] = "Shares Outstanding"
        ws[f'B{row}'] = shares
        apply_style_to_cell(ws[f'A{row}'], get_header_style())
        apply_style_to_cell(ws[f'B{row}'], get_calculation_style())
        ws[f'B{row}'].number_format = config.EXCEL_FORMATTING['number_format']
        row += 1
        
        # Value per Share
        ws[f'A{row}'] = "Value per Share"
        ws[f'B{row}'] = f"=B{row-2}/B{row-1}"
        apply_style_to_cell(ws[f'A{row}'], get_header_style())
        apply_style_to_cell(ws[f'B{row}'], get_formula_style())
        ws[f'B{row}'].number_format = config.EXCEL_FORMATTING['currency_format']
        
        auto_adjust_column_width(ws)
    
    def _create_sensitivity_analysis(self):
        """Create Sensitivity Analysis sheet"""
        ws = self.wb.create_sheet("Sensitivity Analysis")
        
        row = 1
        ws[f'A{row}'] = "Sensitivity Analysis"
        apply_style_to_cell(ws[f'A{row}'], get_header_style())
        row += 2
        
        if self.valuation_analyzer.sensitivity_results:
            # WACC Sensitivity
            if 'wacc' in self.valuation_analyzer.sensitivity_results:
                ws[f'A{row}'] = "WACC Sensitivity"
                apply_style_to_cell(ws[f'A{row}'], get_header_style())
                row += 1
                
                wacc_df = self.valuation_analyzer.sensitivity_results['wacc']
                if not wacc_df.empty:
                    # Headers
                    for col_idx, col_name in enumerate(wacc_df.columns, start=1):
                        ws.cell(row, col_idx, col_name)
                        apply_style_to_cell(ws.cell(row, col_idx), get_header_style())
                    row += 1
                    
                    # Data
                    for _, data_row in wacc_df.iterrows():
                        for col_idx, col_name in enumerate(wacc_df.columns, start=1):
                            value = data_row[col_name]
                            ws.cell(row, col_idx, value)
                            apply_style_to_cell(ws.cell(row, col_idx), get_calculation_style())
                        row += 1
                    row += 1
        
        auto_adjust_column_width(ws)
    
    def _create_scenario_analysis(self):
        """Create Scenario Analysis sheet"""
        ws = self.wb.create_sheet("Scenario Analysis")
        
        row = 1
        ws[f'A{row}'] = "Scenario Analysis"
        apply_style_to_cell(ws[f'A{row}'], get_header_style())
        row += 2
        
        if self.valuation_analyzer.scenario_results:
            scenarios = ['base', 'bull', 'bear']
            ws[f'A{row}'] = "Scenario"
            ws[f'B{row}'] = "Value per Share"
            ws[f'C{row}'] = "Upside/Downside"
            ws[f'D{row}'] = "Recommendation"
            for col in ['A', 'B', 'C', 'D']:
                apply_style_to_cell(ws[f'{col}{row}'], get_header_style())
            row += 1
            
            for scenario in scenarios:
                if scenario in self.valuation_analyzer.scenario_results:
                    data = self.valuation_analyzer.scenario_results[scenario]
                    ws[f'A{row}'] = scenario.capitalize()
                    ws[f'B{row}'] = data.get('value_per_share', 0)
                    ws[f'C{row}'] = data.get('upside_downside_pct', 0) / 100
                    ws[f'D{row}'] = data.get('recommendation', 'N/A')
                    
                    for col in ['A', 'B', 'C', 'D']:
                        if col == 'A' or col == 'D':
                            apply_style_to_cell(ws[f'{col}{row}'], get_header_style())
                        else:
                            apply_style_to_cell(ws[f'{col}{row}'], get_calculation_style())
                    
                    ws[f'B{row}'].number_format = config.EXCEL_FORMATTING['currency_format']
                    ws[f'C{row}'].number_format = config.EXCEL_FORMATTING['percentage_format']
                    row += 1
        
        auto_adjust_column_width(ws)
    
    def _create_relative_valuation(self):
        """Create Relative Valuation sheet"""
        ws = self.wb.create_sheet("Relative Valuation")
        
        row = 1
        ws[f'A{row}'] = "Relative Valuation - Peer Multiples"
        apply_style_to_cell(ws[f'A{row}'], get_header_style())
        row += 2
        
        if self.valuation_analyzer.relative_valuation:
            if 'peer_multiples' in self.valuation_analyzer.relative_valuation:
                peer_df = self.valuation_analyzer.relative_valuation['peer_multiples']
                if not peer_df.empty:
                    # Headers
                    for col_idx, col_name in enumerate(peer_df.columns, start=1):
                        ws.cell(row, col_idx, col_name)
                        apply_style_to_cell(ws.cell(row, col_idx), get_header_style())
                    row += 1
                    
                    # Data
                    for _, data_row in peer_df.iterrows():
                        for col_idx, col_name in enumerate(peer_df.columns, start=1):
                            value = data_row[col_name]
                            ws.cell(row, col_idx, value)
                            apply_style_to_cell(ws.cell(row, col_idx), get_calculation_style())
                        row += 1
        
        auto_adjust_column_width(ws)
    
    def _create_summary(self):
        """Create Summary sheet"""
        ws = self.wb.create_sheet("Summary")
        
        row = 1
        ws[f'A{row}'] = "Valuation Summary"
        apply_style_to_cell(ws[f'A{row}'], get_header_style())
        row += 2
        
        summary_items = [
            ("Company", config.COMPANY_NAME),
            ("Ticker", config.COMPANY_TICKER),
            ("Recommendation", self.final_recommendation.get('recommendation', 'N/A')),
            ("Current Price", self.final_recommendation.get('current_price', 0)),
            ("Target Price", self.final_recommendation.get('target_price', 0)),
            ("Upside/Downside", f"{self.final_recommendation.get('upside_downside_pct', 0):.1f}%"),
        ]
        
        for label, value in summary_items:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            apply_style_to_cell(ws[f'A{row}'], get_header_style())
            apply_style_to_cell(ws[f'B{row}'], get_calculation_style())
            if 'Price' in label:
                ws[f'B{row}'].number_format = config.EXCEL_FORMATTING['currency_format']
            row += 1
        
        auto_adjust_column_width(ws)

