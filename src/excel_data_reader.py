"""
Excel Data Reader
Reads historical financial data from the output Excel file's Historical Financials sheet
This allows users to manually correct data in Excel and have the system use those corrections
"""

import pandas as pd
import os
import sys
from typing import Dict, Optional
from openpyxl import load_workbook

# Add parent directory to path for config import
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config


class ExcelDataReader:
    """Reads historical financial data from the output Excel file"""
    
    def __init__(self, excel_path: Optional[str] = None):
        """
        Initialize Excel data reader
        
        Args:
            excel_path: Path to Excel file (defaults to config.EXCEL_OUTPUT_FILE)
        """
        self.excel_path = excel_path or config.EXCEL_OUTPUT_FILE
    
    def read_historical_financials(self) -> Dict:
        """
        Read historical financial data from Excel file's Historical Financials sheet
        
        Returns:
            Dictionary with 'income_statement', 'balance_sheet', 'cash_flow' DataFrames
            Returns empty DataFrames if file not found or sheet doesn't exist
        """
        result = {
            'income_statement': pd.DataFrame(),
            'balance_sheet': pd.DataFrame(),
            'cash_flow': pd.DataFrame(),
        }
        
        if not os.path.exists(self.excel_path):
            return result
        
        try:
            print(f"  Reading historical data from: {self.excel_path}")
            wb = load_workbook(self.excel_path, data_only=True)
            
            if 'Historical Financials' not in wb.sheetnames:
                print(f"    Warning: 'Historical Financials' sheet not found in Excel file")
                return result
            
            # Read the sheet
            df = pd.read_excel(self.excel_path, sheet_name='Historical Financials', header=None)
            
            # Parse the sheet structure
            result = self._parse_historical_financials_sheet(df)
            
            if not result['income_statement'].empty or not result['balance_sheet'].empty:
                print(f"    ✓ Successfully read historical data from Excel")
            
        except Exception as e:
            print(f"    Warning: Error reading Excel file: {e}")
        
        return result
    
    def _parse_historical_financials_sheet(self, df: pd.DataFrame) -> Dict:
        """
        Parse the Historical Financials sheet structure
        
        Expected structure:
        - "Income Statement (Historical)" header
        - Column A: Period/Date
        - Column B+: Line items (Revenue, EBITDA, Net Income, etc.)
        - "Balance Sheet (Historical)" header
        - Same structure for balance sheet
        - "Cash Flow Statement (Historical)" header
        - Same structure for cash flow
        
        Returns:
            Dictionary with parsed DataFrames
        """
        result = {
            'income_statement': pd.DataFrame(),
            'balance_sheet': pd.DataFrame(),
            'cash_flow': pd.DataFrame(),
        }
        
        if df.empty:
            return result
        
        # Find section headers
        income_start = None
        balance_start = None
        cash_flow_start = None
        
        for idx, row in df.iterrows():
            first_cell = str(row.iloc[0]).lower() if pd.notna(row.iloc[0]) else ""
            
            if 'income statement' in first_cell and 'historical' in first_cell:
                income_start = idx
            elif 'balance sheet' in first_cell and 'historical' in first_cell:
                balance_start = idx
            elif 'cash flow' in first_cell and 'historical' in first_cell:
                cash_flow_start = idx
        
        # Extract each section
        if income_start is not None:
            result['income_statement'] = self._extract_section(
                df, income_start, balance_start if balance_start else cash_flow_start
            )
        
        if balance_start is not None:
            result['balance_sheet'] = self._extract_section(
                df, balance_start, cash_flow_start
            )
        
        if cash_flow_start is not None:
            result['cash_flow'] = self._extract_section(
                df, cash_flow_start, None
            )
        
        return result
    
    def _extract_section(self, df: pd.DataFrame, start_row: int, end_row: Optional[int]) -> pd.DataFrame:
        """
        Extract a financial statement section from the DataFrame
        
        Args:
            df: Full DataFrame
            start_row: Starting row index (section header row)
            end_row: Ending row index (None if last section)
        
        Returns:
            DataFrame with Date column and line item columns
        """
        if start_row is None:
            return pd.DataFrame()
        
        # Skip section header row, then find the header row with column names
        # Usually: section title (row N), empty row (N+1), column headers (N+2), data (N+3+)
        data_start = start_row + 2  # Skip title and empty row
        
        # Determine end row
        if end_row is not None:
            data_end = end_row
        else:
            data_end = len(df)
        
        # Extract the section
        section_df = df.iloc[data_start:data_end].copy()
        
        if section_df.empty or len(section_df) < 2:
            return pd.DataFrame()
        
        # First row should be column headers
        # Set first row as column names
        section_df.columns = section_df.iloc[0]
        section_df = section_df.iloc[1:].reset_index(drop=True)
        
        # Rename first column to "Date" if it's "Period" or similar
        first_col = section_df.columns[0]
        if 'period' in str(first_col).lower() or 'date' in str(first_col).lower():
            section_df = section_df.rename(columns={first_col: 'Date'})
        elif first_col != 'Date':
            # If first column doesn't look like Date, check if it contains dates
            try:
                pd.to_datetime(section_df.iloc[:, 0], errors='raise')
                section_df = section_df.rename(columns={first_col: 'Date'})
            except:
                pass
        
        # Convert Date column to datetime
        if 'Date' in section_df.columns:
            section_df['Date'] = pd.to_datetime(section_df['Date'], errors='coerce')
        
        # Remove empty rows
        section_df = section_df.dropna(how='all')
        
        # Remove columns that are all NaN
        section_df = section_df.dropna(axis=1, how='all')
        
        # Clean column names (remove any extra whitespace)
        section_df.columns = [str(col).strip() if pd.notna(col) else f'Column_{i}' 
                             for i, col in enumerate(section_df.columns)]
        
        # Convert numeric columns (all except Date)
        for col in section_df.columns:
            if col != 'Date':
                section_df[col] = pd.to_numeric(section_df[col], errors='coerce')
        
        return section_df

